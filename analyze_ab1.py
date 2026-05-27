#!/usr/bin/env python3
"""
AB1 Sequencing File Analyzer - F+R Consensus Workflow
从 AB1 文件提取序列，将同一样品的 F+R 合并为共识序列，再与参考序列比对
"""

import os
import re
import json
import argparse
import warnings
from collections import defaultdict
from datetime import datetime

try:
    import numpy as np
except ImportError:
    np = None

try:
    from Bio import SeqIO
    HAS_BIO = True
except ImportError:
    HAS_BIO = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============ 序列处理 ============

COMPLEMENT = str.maketrans('ATGCNRYSWKMBDHVatgcnryswkmbdhv',
                            'TACGNYRSWMKVHBHtacgnyrswmkvhb')

def reverse_complement(seq):
    """返回序列的反向互补 (5'→3')"""
    return seq.translate(COMPLEMENT)[::-1]


def trim_primer_by_prefix(seq, primer, min_match=15):
    """
    从序列开头 trim 引物序列
    使用最长匹配原则，允许少量错配
    """
    if not primer or not seq:
        return seq
    p = primer.upper()
    for length in range(min(len(p), len(seq), 50), min_match - 1, -1):
        prefix = seq[:length].upper()
        # 允许最多2个错配
        mismatches = sum(a != b for a, b in zip(prefix, p[:length]))
        if mismatches <= 2:
            return seq[length:]
    return seq


def hq_trim(seq, qual, min_q=15):
    """
    质量过滤：从两端去除低质量碱基 (Phred < min_q)
    """
    seq = list(seq)
    qual = list(qual)
    while qual and qual[0] < min_q:
        qual.pop(0); seq.pop(0)
    while qual and qual[-1] < min_q:
        qual.pop(); seq.pop()
    return ''.join(seq), qual


def read_ab1(ab1_path):
    """读取 AB1 文件，返回 (序列, 质量分数)"""
    if not HAS_BIO:
        raise ImportError("需要 Biopython: pip install biopython")
    rec = SeqIO.read(ab1_path, 'abi')
    seq = str(rec.seq).upper()
    qual = list(rec.letter_annotations.get('phred_quality', []))
    return seq, qual


# ============ 序列比对 ============

def local_align(query, ref, k=50, step=5):
    """
    简化的局部比对 (sliding window + 动态延伸)
    返回 (best_identity, matched_length, ref_start_pos)
    """
    if not query or not ref:
        return 0.0, 0, 0

    best = (0.0, 0, 0)
    for qs in range(0, max(1, len(query) - k), step):
        q_seg = query[qs:qs + k]
        for rs in range(0, max(1, len(ref) - k), step):
            r_seg = ref[rs:rs + k]
            ids = sum(a == b for a, b in zip(q_seg, r_seg)) / k
            if ids > best[0]:
                ext = min(len(query) - qs, len(ref) - rs)
                if ext < 20:
                    continue
                ids2 = sum(a == b for a, b in zip(query[qs:qs + ext], ref[rs:rs + ext])) / ext
                best = max(best, (ids2, ext, rs), key=lambda x: (x[0], x[1]))
    return best


def find_consensus(f_seq, r_seq):
    """
    尝试 overlap 组装 F + R_RC
    返回 (consensus, method, ov_identity)
    method: 'OVERLAP' / 'CONCAT' / 'F_ONLY'
    """
    if not f_seq:
        return r_seq, 'R_ONLY', 0.0
    if not r_seq:
        return f_seq, 'F_ONLY', 0.0

    best = (0.0, 0, 0)
    for ov in range(15, min(len(f_seq), len(r_seq)) - 5):
        seg_a = f_seq[-(ov + 20):]
        seg_b = r_seq[:ov]
        ids = sum(a == b for a, b in zip(seg_a[-ov:], seg_b)) / ov if ov > 0 else 0
        if ids > best[0]:
            best = (ids, ov, len(f_seq) - ov - 20 + len(r_seq))

    if best[0] >= 0.85:
        return f_seq + r_seq[best[1]:], 'OVERLAP', best[0]
    return f_seq + r_seq, 'CONCAT', best[0]


# ============ 参考序列读取 ============

def read_ref_from_xlsx(xlsx_path, sheet_name='Part',
                        col_name=1, col_seq=4,
                        name_header=None, seq_header=None):
    """
    从 Excel 文件读取参考序列
    col: 1-based 列索引
    返回 {part_name: sequence_str}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    # 自动检测表头行
    header_row = 1
    if name_header or seq_header:
        for r in range(1, min(5, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or '').strip()
                if name_header and name_header.lower() in val.lower():
                    col_name = c
                if seq_header and seq_header.lower() in val.lower():
                    col_seq = c
            header_row = r
            break

    refs = {}
    for row in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(row, col_name).value or '').strip()
        seq_raw = ws.cell(row, col_seq).value
        if not name or not seq_raw:
            continue
        seq = re.sub(r'[^A-Za-z]', '', str(seq_raw)).upper()
        if len(seq) < 10:
            continue
        refs[name] = seq

    return refs


# ============ F+R 文件配对 ============

def pair_ab1_files(ab1_dir, file_list):
    """
    根据文件名将 F 和 R 的 AB1 文件配对
    文件命名规范: <ID>_[pScbbXX-F].ab1 / <ID>_[pScbbXX-R].ab1
    或: <ID>_[F].ab1 / <ID>_[R].ab1
    返回 {sample_id: {'F': path, 'R': path}}
    """
    pairs = {}
    for fname in file_list:
        fpath = os.path.join(ab1_dir, fname)
        name_lower = fname.lower()

        # 提取 sample_id
        # 尝试模式1: <num>_<id>_[F/R].ab1
        m = re.match(r'(\d+_\d+_[^(]+)\[([^\]]+)\]', fname)
        if m:
            sid = m.group(1).strip()
            direction = 'F' if 'f' in m.group(2).lower() else 'R'
        else:
            # 模式2: <id>_F.ab1 / <id>_R.ab1
            m = re.match(r'(.+?)_-?[FR]\.ab1', fname, re.IGNORECASE)
            if m:
                sid = m.group(1).strip()
                direction = 'F' if name_lower.endswith('_f.ab1') else 'R'
            else:
                continue

        if sid not in pairs:
            pairs[sid] = {}
        pairs[sid][direction] = fpath

    return pairs


def extract_sample_id(filename):
    """从文件名提取原始样本ID (如 T34-2-G5)"""
    # 尝试从括号内提取: (T34-2-G5)_
    m = re.search(r'\(([^)]+)\)', filename)
    if m:
        return m.group(1).strip()
    # 或从编号提取: 0049_33326052101663_
    m = re.search(r'\d{4}_\d+_(.+)_\[', filename)
    if m:
        return m.group(1).strip()
    return os.path.splitext(filename)[0]


# ============ 批量分析 ============

def analyze_batch(ab1_dir, refs,
                  primer_f=None,
                  min_identity=0.70,
                  xlsx_info=None,
                  verbose=True):
    """
    批量分析 AB1 文件，F+R 合并共识后与参考序列比对
    """
    ab1_files = sorted([f for f in os.listdir(ab1_dir)
                        if f.lower().endswith('.ab1')])

    # 配对 F+R
    pairs = pair_ab1_files(ab1_dir, ab1_files)

    results = []

    for sample_id, files in sorted(pairs.items()):
        if verbose:
            print(f"处理: {sample_id} ..." , end=' ', flush=True)

        f_path = files.get('F')
        r_path = files.get('R')

        row = {
            'sample_id': sample_id,
            'file_F': os.path.basename(f_path) if f_path else '',
            'file_R': os.path.basename(r_path) if r_path else '',
            'assembly': 'NONE',
            'f_len': 0, 'r_len': 0,
            'consensus_len': 0,
            'best_match': '',
            'best_identity': 0.0,
            'best_coverage': 0.0,
            'target_part': '',
            'status': '❌ 无文件',
            'f_start': '', 'r_start': '',
        }

        if not f_path:
            results.append(row)
            if verbose: print("❌ 缺少F文件")
            continue

        try:
            # === 读取 F ===
            f_seq, f_qual = read_ab1(f_path)
            f_hq, f_qual_hq = hq_trim(f_seq, f_qual)
            f_clean = trim_primer_by_prefix(f_hq, primer_f) if primer_f else f_hq

            # === 读取 R ===
            r_seq, r_qual = '', []
            r_rc_clean = ''
            if r_path and os.path.exists(r_path):
                r_seq, r_qual = read_ab1(r_path)
                r_hq, _ = hq_trim(r_seq, r_qual)
                r_rc = reverse_complement(r_hq)
                r_rc_clean = trim_primer_by_prefix(r_rc, primer_f) if primer_f else r_rc

            # === 记录清洗后序列 ===
            row['f_len'] = len(f_clean)
            row['r_len'] = len(r_rc_clean)
            row['f_start'] = f_clean[:30]
            row['r_start'] = r_rc_clean[:30] if r_rc_clean else ''

            # === 组装共识序列 ===
            consensus, method, ov_ids = find_consensus(f_clean, r_rc_clean)
            row['assembly'] = method
            row['consensus_len'] = len(consensus)

            if verbose:
                print(f"[{method}] ", end='', flush=True)

            # === 与所有参考序列比对 ===
            best_overall = (0.0, 0, '')
            for ref_name, ref_seq in refs.items():
                ids, aln_len, r_pos = local_align(consensus, ref_seq)
                coverage = aln_len / len(ref_seq) if ref_seq else 0
                if ids > best_overall[0] or (ids == best_overall[0] and coverage > best_overall[1] / len(refs.get(best_overall[2], 'X'))):
                    best_overall = (ids, aln_len, ref_name, coverage, r_pos)

            if len(best_overall) == 5:
                ids, aln_len, ref_name, coverage, r_pos = best_overall
                row['best_match'] = ref_name
                row['best_identity'] = ids
                row['best_coverage'] = coverage
                row['status'] = _judge(ids, coverage)
            else:
                row['best_match'] = best_overall[2]
                row['best_identity'] = best_overall[0]
                row['best_coverage'] = 0.0
                row['status'] = _judge(best_overall[0], 0)

            # === 标注目标 Part ===
            if xlsx_info:
                row['target_part'] = xlsx_info.get(sample_id, '')

            if verbose:
                print(f"{row['best_match']} ({row['best_identity']*100:.1f}%)")

        except Exception as e:
            row['status'] = f'❌ 错误: {e}'
            if verbose:
                print(f"❌ {e}")

        results.append(row)

    return results


def _judge(identity, coverage):
    """判定比对结果"""
    if identity >= 0.95 and coverage >= 0.70:
        return '✅ 正确'
    elif identity >= 0.80 and coverage >= 0.50:
        return '⚠️ 部分'
    elif identity >= 0.70:
        return '⚠️ 低'
    else:
        return '❌ 不匹配'


# ============ Excel 报告生成 ============

def generate_excel(results, output_path):
    """生成格式化的 Excel 比对报告"""
    if not HAS_OPENPYXL:
        raise ImportError("需要 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "比对结果"

    # 样式
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "克隆ID", "目标Part", "F文件名", "R文件名",
        "组装方式", "F清洗长", "R_RC清洗长", "共识长",
        "最佳匹配", "Identity(%)", "覆盖度(%)",
        "判定", "F序列开头(30bp)", "R序列开头(30bp)"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for r_idx, row in enumerate(results, 2):
        status = row['status']
        if '✅' in status:
            row_fill = pass_fill
        elif '⚠️' in status:
            row_fill = warn_fill
        else:
            row_fill = fail_fill

        ws.append([
            row['sample_id'],
            row['target_part'],
            row['file_F'],
            row['file_R'],
            row['assembly'],
            row['f_len'],
            row['r_len'],
            row['consensus_len'],
            row['best_match'],
            f"{row['best_identity']*100:.1f}" if row['best_identity'] else "-",
            f"{row['best_coverage']*100:.1f}" if row['best_coverage'] else "-",
            status,
            row['f_start'],
            row['r_start'],
        ])

        for cell in ws[r_idx]:
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = border

    # 列宽
    col_widths = [14, 12, 38, 38, 10, 9, 11, 9, 14, 10, 10, 10, 32, 32]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    wb.save(output_path)
    print(f"报告已保存: {output_path}")


# ============ CLI ============

def main():
    parser = argparse.ArgumentParser(
        description='AB1 F+R 共识序列比对分析工具')
    parser.add_argument('--ab1-dir', '-d', required=True,
                        help='AB1 文件目录')
    parser.add_argument('--ref-xlsx', '-x', required=True,
                        help='参考序列 Excel 文件')
    parser.add_argument('--ref-sheet', '-s', default='Part',
                        help='Excel 工作表名称 (默认: Part)')
    parser.add_argument('--ref-col-name', type=int, default=1,
                        help='Part 名称列 (1-based, 默认: 1)')
    parser.add_argument('--ref-col-seq', type=int, default=4,
                        help='序列列 (1-based, 默认: 4)')
    parser.add_argument('--primer-f', '-p',
                        help='正向引物序列 (用于 trimming)')
    parser.add_argument('--min-identity', '-i', type=float, default=0.70,
                        help='最小 identity (默认: 0.70)')
    parser.add_argument('--output', '-o', default='ab1_consensus_report.xlsx',
                        help='输出报告路径 (默认: ab1_consensus_report.xlsx)')
    args = parser.parse_args()

    if not os.path.isdir(args.ab1_dir):
        print(f"❌ 目录不存在: {args.ab1_dir}")
        return
    if not os.path.exists(args.ref_xlsx):
        print(f"❌ Excel 文件不存在: {args.ref_xlsx}")
        return

    print("=" * 60)
    print("AB1 F+R 共识序列比对分析")
    print("=" * 60)

    # 读取参考序列
    print(f"读取参考序列: {args.ref_xlsx} / {args.ref_sheet}")
    refs = read_ref_from_xlsx(
        args.ref_xlsx, args.ref_sheet,
        col_name=args.ref_col_name, col_seq=args.ref_col_seq)
    print(f"共读取 {len(refs)} 个参考序列")

    # 批量分析
    print(f"\n开始分析 {args.ab1_dir} ...")
    results = analyze_batch(
        args.ab1_dir, refs,
        primer_f=args.primer_f,
        min_identity=args.min_identity)

    # 生成报告
    print(f"\n生成 Excel 报告 ...")
    generate_excel(results, args.output)

    # 统计
    cnt = defaultdict(int)
    for r in results:
        cnt[r['status']] += 1
    print("\n统计:")
    for k, v in sorted(cnt.items()):
        print(f"  {k}: {v}")
    print(f"\n总计: {len(results)} 个克隆")
    print("=" * 60)


if __name__ == '__main__':
    main()
